#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ============================================================================
# jar_report.py
#   指定ディレクトリ配下の JAR を走査し、Maven 座標・取得 URL・ライブラリ概要
#   を解決して「美しい」Excel (.xlsx) に出力するレポート機能。
#
#   jar_manager.py の --mode report から呼び出されるほか、単体実行も可能。
#     python3 jar_report.py --scan-dir /opt/app/lib --output report.xlsx
#     python3 jar_report.py --scan-dir /opt/app/lib --output report.xlsx --online
#
#   座標の解決優先順位 (groupId / artifactId / version):
#     1. JAR 内 META-INF/maven/**/pom.properties          (最も確実)
#     2. JAR 内 META-INF/maven/**/pom.xml                  (parent 継承も考慮)
#     3. MANIFEST.MF (Implementation-*/Bundle-*)           (version 等を補完)
#     4. 既知アーティファクト → groupId マッピング (offline)
#     5. Maven Central への SHA-1 / 座標照合 (--online 時のみ)
#     6. ファイル名からの推定 (artifactId / version / classifier)
#   groupId や version が欠落していても、残るファイル名 + 上記手段で推測する。
#
#   ライブラリ概要は pom.xml <name>/<description>、MANIFEST の title/description、
#   および内蔵の既知説明マップから構築する。
#
#   依存: openpyxl (Excel 出力用)。未導入時は明確なエラーで案内する。
#         走査・座標解決部分は Python 標準ライブラリのみで動作する。
# ============================================================================
from __future__ import annotations

import argparse
import datetime
import json
import logging
import os
import sys
import urllib.error
import urllib.parse
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Dict, List, Optional

# jar_manager.py の既存ロジックを再利用 (単一の判定基準を保つ)。
try:
    from jar_manager import (
        DEFAULT_REPOS,
        parse_jar_filename,
        read_pom_properties,
        compute_hash,
        http_fetch_text,
        http_exists,
        setup_logging,
    )
except ImportError:  # 単体配置された場合のフォールバック (最小限)
    import re as _re
    DEFAULT_REPOS = ["https://repo1.maven.org/maven2"]
    _VER_START_RE = _re.compile(r"-(?=\d)")

    def parse_jar_filename(filename):  # type: ignore
        base = os.path.basename(filename)
        stem = base[:-4] if base.lower().endswith(".jar") else base
        m = _VER_START_RE.search(stem)
        if not m:
            return (stem, "UNKNOWN", "")
        idx = m.start()
        return (stem[:idx], stem[idx + 1:].split("-")[0], "")

    def read_pom_properties(jar_path):  # type: ignore
        return None

    def compute_hash(path, algo):  # type: ignore
        import hashlib
        h = hashlib.new(algo)
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 256), b""):
                h.update(chunk)
        return h.hexdigest()

    def http_fetch_text(url, timeout):  # type: ignore
        try:
            import urllib.request
            req = urllib.request.Request(
                url, headers={"User-Agent": "jar_manager/1.0"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                if 200 <= resp.status < 300:
                    return resp.read().decode("utf-8", "replace")
        except Exception:
            return None
        return None

    def http_exists(url, timeout, logger):  # type: ignore
        return False

    def setup_logging(log_file):  # type: ignore
        logging.basicConfig(level=logging.INFO)
        return logging.getLogger("jar_report")


# ----------------------------------------------------------------------------
# 既知アーティファクト → groupId マッピング (offline 推定用)
#   pom.properties も pom.xml も無い JAR で groupId を補うための最小限の辞書。
#   ここに無いものは --online (Maven Central 照合) か UNKNOWN で対応する。
# ----------------------------------------------------------------------------
KNOWN_GROUPS: Dict[str, str] = {
    "log4j-core": "org.apache.logging.log4j",
    "log4j-api": "org.apache.logging.log4j",
    "log4j-1.2-api": "org.apache.logging.log4j",
    "jackson-databind": "com.fasterxml.jackson.core",
    "jackson-core": "com.fasterxml.jackson.core",
    "jackson-annotations": "com.fasterxml.jackson.core",
    "spring-core": "org.springframework",
    "spring-context": "org.springframework",
    "spring-beans": "org.springframework",
    "spring-web": "org.springframework",
    "spring-webmvc": "org.springframework",
    "guava": "com.google.guava",
    "gson": "com.google.code.gson",
    "commons-lang3": "org.apache.commons",
    "commons-text": "org.apache.commons",
    "commons-collections4": "org.apache.commons",
    "commons-io": "commons-io",
    "commons-codec": "commons-codec",
    "commons-logging": "commons-logging",
    "slf4j-api": "org.slf4j",
    "slf4j-simple": "org.slf4j",
    "logback-core": "ch.qos.logback",
    "logback-classic": "ch.qos.logback",
    "junit": "junit",
    "junit-jupiter-api": "org.junit.jupiter",
    "hibernate-core": "org.hibernate",
    "mysql-connector-j": "com.mysql",
    "postgresql": "org.postgresql",
    "httpclient": "org.apache.httpcomponents",
    "httpcore": "org.apache.httpcomponents",
    "snakeyaml": "org.yaml",
    "bcprov-jdk18on": "org.bouncycastle",
    "netty-all": "io.netty",
    "protobuf-java": "com.google.protobuf",
    "lombok": "org.projectlombok",
}

# 既知ライブラリの日本語概要 (offline で概要欄を充実させるための補助)。
KNOWN_DESCRIPTIONS: Dict[str, str] = {
    "log4j-core": "Apache Log4j 2 のコア実装。高性能な Java ロギングフレームワーク。",
    "log4j-api": "Apache Log4j 2 の API。ロギングのファサードを提供する。",
    "jackson-databind": "Jackson のデータバインディング。JSON と Java オブジェクトの相互変換。",
    "jackson-core": "Jackson のコア。ストリーミング JSON パーサ/ジェネレータ。",
    "spring-core": "Spring Framework のコアユーティリティと DI 基盤。",
    "guava": "Google Guava。コレクション・キャッシュ・並行処理等のコアライブラリ。",
    "gson": "Google Gson。JSON と Java オブジェクトの相互変換ライブラリ。",
    "commons-lang3": "Apache Commons Lang。java.lang を補完するユーティリティ群。",
    "commons-io": "Apache Commons IO。入出力処理のユーティリティ群。",
    "slf4j-api": "SLF4J API。ロギング実装を抽象化するファサード。",
    "logback-classic": "Logback。SLF4J のネイティブ実装ロギングフレームワーク。",
    "junit": "JUnit。Java の単体テストフレームワーク。",
    "snakeyaml": "SnakeYAML。Java 向け YAML パーサ/エミッタ。",
    "httpclient": "Apache HttpClient。HTTP 通信クライアントライブラリ。",
}

# リポジトリ base URL → 表示名。
REPO_DISPLAY_NAMES: Dict[str, str] = {
    "https://repo1.maven.org/maven2": "Maven Central",
    "https://repository.apache.org/content/repositories/releases": "Apache Releases",
    "https://repo.spring.io/release": "Spring",
    "https://repository.jboss.org/nexus/content/repositories/releases": "JBoss",
    "https://oss.sonatype.org/content/repositories/releases": "Sonatype Releases",
}

UNKNOWN = "UNKNOWN"


# ----------------------------------------------------------------------------
# レポート 1 行分のデータ
# ----------------------------------------------------------------------------
@dataclass
class JarReport:
    fileName: str = ""
    groupId: str = ""
    artifactId: str = ""
    version: str = ""
    classifier: str = ""
    packaging: str = "jar"
    libName: str = ""       # ライブラリ名 (人が読む名称)
    description: str = ""   # 概要
    vendor: str = ""        # 提供元/ベンダー
    license: str = ""       # ライセンス
    homepage: str = ""      # プロジェクト URL
    coordSource: str = ""   # 座標の判定元 (pom.properties 等)
    inferred: bool = False  # 座標のいずれかを推測で補ったか
    downloadUrl: str = ""   # 取得 URL
    repoName: str = ""      # 取得元リポジトリ表示名
    urlVerified: str = ""   # URL 実在確認結果 (online 時)
    sizeBytes: int = 0
    sha1: str = ""
    targetDir: str = ""     # 実際の配置ディレクトリ
    note: str = ""          # 備考


# ----------------------------------------------------------------------------
# JAR 内メタデータ抽出
# ----------------------------------------------------------------------------
def read_manifest(jar_path: str) -> Dict[str, str]:
    """META-INF/MANIFEST.MF の主属性を dict で返す (継続行 unfold 対応)。"""
    try:
        with zipfile.ZipFile(jar_path) as zf:
            if "META-INF/MANIFEST.MF" not in zf.namelist():
                return {}
            raw = zf.read("META-INF/MANIFEST.MF").decode("utf-8", "replace")
    except (zipfile.BadZipFile, OSError, KeyError):
        return {}
    # 行頭スペースは直前行の継続 (RFC 822 スタイル)。
    unfolded: List[str] = []
    for line in raw.splitlines():
        if line[:1] == " " and unfolded:
            unfolded[-1] += line[1:]
        else:
            unfolded.append(line)
    attrs: Dict[str, str] = {}
    for line in unfolded:
        if ":" in line:
            k, _, v = line.partition(":")
            k = k.strip()
            if k:
                attrs[k] = v.strip()
    return attrs


def _local(tag: str) -> str:
    """XML タグから名前空間を除いたローカル名を返す。"""
    return tag.rsplit("}", 1)[-1]


def read_embedded_pom(jar_path: str) -> Dict[str, str]:
    """META-INF/maven/**/pom.xml から name/description/url/license と、
    (pom.properties が無い場合の) groupId/artifactId/version を抽出する。
    parent からの groupId/version 継承も考慮する。
    """
    try:
        with zipfile.ZipFile(jar_path) as zf:
            cands = [n for n in zf.namelist()
                     if n.startswith("META-INF/maven/") and n.endswith("pom.xml")]
            if not cands:
                return {}
            text = zf.read(sorted(cands)[0]).decode("utf-8", "replace")
    except (zipfile.BadZipFile, OSError, KeyError):
        return {}
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return {}

    res: Dict[str, str] = {}
    parent: Dict[str, str] = {}
    licenses: List[str] = []
    for child in list(root):
        t = _local(child.tag)
        if t in ("name", "description", "url") and child.text and child.text.strip():
            res.setdefault(t, child.text.strip())
        elif t in ("groupId", "artifactId", "version") and child.text:
            res[t] = child.text.strip()
        elif t == "parent":
            for pc in list(child):
                pt = _local(pc.tag)
                if pt in ("groupId", "version") and pc.text:
                    parent[pt] = pc.text.strip()
        elif t == "licenses":
            for lic in list(child):
                for lc in list(lic):
                    if _local(lc.tag) == "name" and lc.text and lc.text.strip():
                        licenses.append(lc.text.strip())
    # groupId / version は parent から継承され得る。
    if "groupId" not in res and parent.get("groupId"):
        res["groupId"] = parent["groupId"]
    if "version" not in res and parent.get("version"):
        res["version"] = parent["version"]
    if licenses:
        res["license"] = ", ".join(dict.fromkeys(licenses))
    return res


# ----------------------------------------------------------------------------
# Maven Central 照合 (--online)
# ----------------------------------------------------------------------------
_SEARCH_URL = "https://search.maven.org/solrsearch/select"


def _mc_query(query: str, timeout: int, logger: logging.Logger
              ) -> Optional[Dict[str, str]]:
    url = _SEARCH_URL + "?" + urllib.parse.urlencode(
        {"q": query, "rows": "5", "wt": "json"})
    txt = http_fetch_text(url, timeout)
    if not txt:
        return None
    try:
        data = json.loads(txt)
        docs = data.get("response", {}).get("docs", [])
    except (ValueError, AttributeError):
        return None
    if not docs:
        return None
    d = docs[0]
    g = d.get("g") or d.get("groupId")
    a = d.get("a") or d.get("artifactId")
    v = d.get("v") or d.get("latestVersion") or d.get("version")
    if not (g and a):
        return None
    return {"groupId": g, "artifactId": a, "version": v or ""}


def maven_central_by_sha1(sha1: str, timeout: int,
                          logger: logging.Logger) -> Optional[Dict[str, str]]:
    """SHA-1 で Maven Central を照合し、正確な座標を得る (最も確実な online 手段)。"""
    return _mc_query('1:"%s"' % sha1, timeout, logger)


def maven_central_by_ga(artifact: str, version: str, timeout: int,
                        logger: logging.Logger) -> Optional[Dict[str, str]]:
    """artifactId (+ version) で Maven Central を照合し groupId 等を得る。"""
    q = 'a:"%s"' % artifact
    if version and version != UNKNOWN:
        q += ' AND v:"%s"' % version
    return _mc_query(q, timeout, logger)


def fetch_pom_metadata(repo_base: str, group: str, artifact: str, version: str,
                       timeout: int) -> Dict[str, str]:
    """リモートの pom を取得し name/description/url/license を抽出する。"""
    gp = group.replace(".", "/")
    url = "%s/%s/%s/%s/%s-%s.pom" % (
        repo_base.rstrip("/"), gp, artifact, version, artifact, version)
    txt = http_fetch_text(url, timeout)
    if not txt:
        return {}
    try:
        root = ET.fromstring(txt)
    except ET.ParseError:
        return {}
    res: Dict[str, str] = {}
    licenses: List[str] = []
    for child in list(root):
        t = _local(child.tag)
        if t in ("name", "description", "url") and child.text and child.text.strip():
            res.setdefault(t, child.text.strip())
        elif t == "licenses":
            for lic in list(child):
                for lc in list(lic):
                    if _local(lc.tag) == "name" and lc.text and lc.text.strip():
                        licenses.append(lc.text.strip())
    if licenses:
        res["license"] = ", ".join(dict.fromkeys(licenses))
    return res


# ----------------------------------------------------------------------------
# 取得 URL 解決
# ----------------------------------------------------------------------------
def repo_display(base: str) -> str:
    return REPO_DISPLAY_NAMES.get(base.rstrip("/"), base)


def build_download_url(repo_base: str, group: str, artifact: str,
                       version: str, classifier: str) -> str:
    gp = group.replace(".", "/")
    if classifier:
        fname = "%s-%s-%s.jar" % (artifact, version, classifier)
    else:
        fname = "%s-%s.jar" % (artifact, version)
    return "%s/%s/%s/%s/%s" % (repo_base.rstrip("/"), gp, artifact, version, fname)


def resolve_download_url(group: str, artifact: str, version: str,
                         classifier: str, repos: List[str], online: bool,
                         timeout: int, logger: logging.Logger):
    """(url, repoName, verified) を返す。

    * 座標が未確定 (UNKNOWN) の場合は URL 生成不可を返す。
    * online 時は各リポジトリに HEAD で存在確認し、実在した最初のものを採用。
    * offline 時は先頭リポジトリ (通常 Maven Central) の候補 URL を返す。
    """
    if group == UNKNOWN or version == UNKNOWN or not artifact:
        return ("", "", "座標未確定のため生成不可")
    if online:
        for repo in repos:
            url = build_download_url(repo, group, artifact, version, classifier)
            try:
                if http_exists(url, timeout, logger):
                    return (url, repo_display(repo), "実在確認 OK")
            except Exception:  # noqa: BLE001  ネットワーク例外は握りつぶし継続
                continue
        # どこにも無い: 候補として Central を提示
        url = build_download_url(repos[0], group, artifact, version, classifier)
        return (url, repo_display(repos[0]), "全リポジトリで未検出")
    url = build_download_url(repos[0], group, artifact, version, classifier)
    return (url, repo_display(repos[0]) + " (候補)", "未確認 (offline)")


# ----------------------------------------------------------------------------
# 1 JAR → JarReport 構築
# ----------------------------------------------------------------------------
def build_report_for_jar(full_path: str, name: str, online: bool,
                         repos: List[str], timeout: int,
                         logger: logging.Logger) -> JarReport:
    art_f, ver_f, cls_f = parse_jar_filename(name)
    props = read_pom_properties(full_path)
    manifest = read_manifest(full_path)
    pom = read_embedded_pom(full_path)

    group = artifact = version = ""
    sources: List[str] = []
    inferred = False
    notes: List[str] = []

    # --- 座標: pom.properties > pom.xml ---
    if props:
        group = props.get("groupId", "")
        artifact = props.get("artifactId", "") or art_f
        version = props.get("version", "") or ver_f
        sources.append("pom.properties")
    else:
        artifact = pom.get("artifactId", "") or art_f
        version = pom.get("version", "") or ver_f
        group = pom.get("groupId", "")
        if group or pom.get("artifactId"):
            sources.append("pom.xml")

    # --- version 補完: MANIFEST ---
    if not version or version == UNKNOWN:
        mv = (manifest.get("Implementation-Version")
              or manifest.get("Bundle-Version"))
        if mv:
            version = mv.split(".v")[0]  # OSGi の .vYYYYMMDD 修飾を除去
            inferred = True
            if "MANIFEST" not in sources:
                sources.append("MANIFEST")

    # --- groupId 補完: 既知マッピング (offline) ---
    if not group:
        guess = KNOWN_GROUPS.get(artifact)
        if guess:
            group = guess
            inferred = True
            sources.append("既知マッピング")

    # --- online: SHA-1 照合 (最優先の確実な補完) ---
    sha1 = ""
    try:
        sha1 = compute_hash(full_path, "sha1")
    except OSError:
        pass

    if online and (not group or version in ("", UNKNOWN)) and sha1:
        doc = maven_central_by_sha1(sha1, timeout, logger)
        if doc:
            group = doc["groupId"]
            artifact = doc["artifactId"] or artifact
            if doc.get("version"):
                version = doc["version"]
            inferred = True
            sources.append("MavenCentral(sha1)")
    if online and not group and artifact:
        doc = maven_central_by_ga(artifact, version, timeout, logger)
        if doc:
            group = doc["groupId"]
            if doc.get("version") and version in ("", UNKNOWN):
                version = doc["version"]
            inferred = True
            sources.append("MavenCentral(座標)")

    if not group:
        group = UNKNOWN
        notes.append("groupId を特定できず。手動補正を推奨。")
    if not version:
        version = UNKNOWN
        notes.append("version を特定できず。")
    if artifact and (art_f != artifact and not props and not pom.get("artifactId")):
        inferred = True

    # --- ライブラリ名 / 概要 / ベンダー / ライセンス / URL ---
    lib_name = (pom.get("name")
                or manifest.get("Implementation-Title")
                or manifest.get("Bundle-Name")
                or artifact)
    description = (pom.get("description")
                   or manifest.get("Bundle-Description")
                   or KNOWN_DESCRIPTIONS.get(artifact, ""))
    vendor = (manifest.get("Implementation-Vendor")
              or manifest.get("Bundle-Vendor")
              or manifest.get("Implementation-Vendor-Id")
              or "")
    license_ = (pom.get("license")
                or manifest.get("Bundle-License", ""))
    homepage = (pom.get("url")
                or manifest.get("Bundle-DocURL")
                or manifest.get("Implementation-URL")
                or "")

    # --- online: 説明が空なら pom を取得して補完 ---
    if online and (not description or not license_) \
            and group != UNKNOWN and version != UNKNOWN:
        meta = fetch_pom_metadata(repos[0], group, artifact, version, timeout)
        if meta:
            description = description or meta.get("description", "")
            lib_name = lib_name or meta.get("name", "")
            homepage = homepage or meta.get("url", "")
            license_ = license_ or meta.get("license", "")

    # --- 取得 URL ---
    url, repo_name, verified = resolve_download_url(
        group, artifact, version, cls_f, repos, online, timeout, logger)

    size = 0
    try:
        size = os.path.getsize(full_path)
    except OSError:
        pass

    return JarReport(
        fileName=name,
        groupId=group,
        artifactId=artifact,
        version=version,
        classifier=cls_f,
        packaging="jar",
        libName=lib_name,
        description=description,
        vendor=vendor,
        license=license_,
        homepage=homepage,
        coordSource=" + ".join(dict.fromkeys(sources)) or "ファイル名推定",
        inferred=inferred,
        downloadUrl=url,
        repoName=repo_name,
        urlVerified=verified,
        sizeBytes=size,
        sha1=sha1,
        targetDir=os.path.dirname(full_path),
        note=" / ".join(notes),
    )


# ----------------------------------------------------------------------------
# ディレクトリ走査
# ----------------------------------------------------------------------------
def scan_jars(scan_dir: str, include_old: bool, online: bool,
              repos: List[str], timeout: int,
              logger: logging.Logger) -> List[JarReport]:
    if not os.path.isdir(scan_dir):
        raise ValueError("scan-dir が存在しません: %s" % scan_dir)
    reports: List[JarReport] = []
    for root, dirs, files in os.walk(scan_dir):
        if not include_old and "old" in dirs:
            dirs[:] = [d for d in dirs if d != "old"]
        for fn in sorted(files):
            if not fn.lower().endswith(".jar"):
                continue
            full = os.path.join(root, fn)
            logger.info("解析中: %s", full)
            reports.append(build_report_for_jar(
                full, fn, online, repos, timeout, logger))
    reports.sort(key=lambda r: (r.groupId == UNKNOWN, r.groupId,
                                r.artifactId, r.version))
    return reports


# ----------------------------------------------------------------------------
# Excel 出力 (openpyxl)
# ----------------------------------------------------------------------------
# 列定義: (見出し, JarReport 属性, 幅, 折り返し)
_COLUMNS = [
    ("No.", None, 5, False),
    ("ファイル名", "fileName", 32, False),
    ("groupId", "groupId", 26, False),
    ("artifactId", "artifactId", 22, False),
    ("version", "version", 14, False),
    ("classifier", "classifier", 12, False),
    ("packaging", "packaging", 10, False),
    ("ライブラリ名", "libName", 26, True),
    ("概要", "description", 46, True),
    ("提供元/ベンダー", "vendor", 24, True),
    ("ライセンス", "license", 20, True),
    ("プロジェクト URL", "homepage", 30, False),
    ("座標の判定元", "coordSource", 22, True),
    ("推定", "inferred", 7, False),
    ("取得 URL", "downloadUrl", 60, False),
    ("取得元リポジトリ", "repoName", 20, False),
    ("URL 確認", "urlVerified", 18, False),
    ("サイズ", "sizeBytes", 12, False),
    ("SHA-1", "sha1", 44, False),
    ("配置ディレクトリ", "targetDir", 30, False),
    ("備考", "note", 30, True),
]


def _human_size(n: int) -> str:
    if not n:
        return ""
    units = ["B", "KB", "MB", "GB"]
    f = float(n)
    for u in units:
        if f < 1024 or u == units[-1]:
            return ("%d %s" % (int(f), u)) if u == "B" else ("%.1f %s" % (f, u))
        f /= 1024
    return str(n)


def write_xlsx(reports: List[JarReport], output: str, scan_dir: str,
               online: bool, logger: logging.Logger) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "Excel 出力には openpyxl が必要です。'pip install openpyxl' で導入して"
            "ください (または python3 -m pip install --user openpyxl)。")

    # --- テーマ色 ---
    C_HEADER_BG = "1F4E78"   # 濃紺
    C_HEADER_FG = "FFFFFF"
    C_TITLE_BG = "2E75B6"
    C_BAND = "EAF1FB"        # 淡い青 (縞)
    C_INFERRED = "FFF2CC"    # 推定行の淡い黄
    C_UNKNOWN = "FCE4D6"     # UNKNOWN の淡い橙
    C_META = "D9E1F2"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "JAR 一覧"

    header_font = Font(name="Meiryo", bold=True, color=C_HEADER_FG, size=10)
    title_font = Font(name="Meiryo", bold=True, color="FFFFFF", size=16)
    meta_font = Font(name="Meiryo", size=9, color="44546A")
    cell_font = Font(name="Meiryo", size=9)
    link_font = Font(name="Meiryo", size=9, color="0563C1", underline="single")

    ncols = len(_COLUMNS)
    last_col = get_column_letter(ncols)

    # --- タイトル行 ---
    ws.merge_cells("A1:%s1" % last_col)
    tcell = ws["A1"]
    tcell.value = "JAR インベントリ・レポート"
    tcell.font = title_font
    tcell.fill = PatternFill("solid", fgColor=C_TITLE_BG)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # --- メタ情報行 ---
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total = len(reports)
    n_unknown = sum(1 for r in reports if r.groupId == UNKNOWN)
    n_inferred = sum(1 for r in reports if r.inferred)
    meta = ("生成日時: %s   |   走査対象: %s   |   総数: %d 件   |   "
            "groupId 未確定: %d 件   |   推定補完: %d 件   |   モード: %s"
            % (now, scan_dir, total, n_unknown, n_inferred,
               "online (Maven Central 照合)" if online else "offline"))
    ws.merge_cells("A2:%s2" % last_col)
    mcell = ws["A2"]
    mcell.value = meta
    mcell.font = meta_font
    mcell.fill = PatternFill("solid", fgColor=C_META)
    mcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    header_row = 3
    data_start = header_row + 1

    # --- ヘッダ ---
    for ci, (label, _attr, width, _wrap) in enumerate(_COLUMNS, 1):
        c = ws.cell(row=header_row, column=ci, value=label)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[header_row].height = 26

    # --- データ ---
    for ri, r in enumerate(reports):
        row = data_start + ri
        is_unknown = r.groupId == UNKNOWN
        if is_unknown:
            row_fill = PatternFill("solid", fgColor=C_UNKNOWN)
        elif r.inferred:
            row_fill = PatternFill("solid", fgColor=C_INFERRED)
        elif ri % 2 == 1:
            row_fill = PatternFill("solid", fgColor=C_BAND)
        else:
            row_fill = None

        for ci, (label, attr, _width, wrap) in enumerate(_COLUMNS, 1):
            if attr is None:
                value = ri + 1
            elif attr == "inferred":
                value = "推定" if r.inferred else ""
            elif attr == "sizeBytes":
                value = _human_size(r.sizeBytes)
            else:
                value = getattr(r, attr)
            c = ws.cell(row=row, column=ci, value=value)
            c.font = cell_font
            c.border = border
            c.alignment = Alignment(
                horizontal="center" if attr in (None, "inferred") else "left",
                vertical="top",
                wrap_text=wrap)
            if row_fill is not None:
                c.fill = row_fill
            # ハイパーリンク化
            if attr == "downloadUrl" and value:
                c.hyperlink = value
                c.font = link_font
            elif attr == "homepage" and value and value.startswith("http"):
                c.hyperlink = value
                c.font = link_font

    # --- オートフィルタ + ウィンドウ枠固定 ---
    ws.auto_filter.ref = "A%d:%s%d" % (header_row, last_col,
                                       max(header_row, data_start + total - 1))
    ws.freeze_panes = "C%d" % data_start  # ファイル名まで固定表示

    # --- 凡例シート ---
    _write_legend_sheet(wb, C_HEADER_BG, C_HEADER_FG, C_INFERRED, C_UNKNOWN,
                        C_BAND, border)

    # 一時ファイル経由で atomic に保存。
    tmp = output + ".tmp"
    wb.save(tmp)
    os.replace(tmp, output)
    logger.info("Excel 出力完了: %s (%d 件)", output, total)


def _write_legend_sheet(wb, c_header_bg, c_header_fg, c_inferred, c_unknown,
                        c_band, border) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("凡例")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    title_font = Font(name="Meiryo", bold=True, size=14, color="1F4E78")
    head_font = Font(name="Meiryo", bold=True, color=c_header_fg, size=10)
    body_font = Font(name="Meiryo", size=9)

    ws["A1"] = "凡例・カラム説明"
    ws["A1"].font = title_font

    ws["A3"] = "行の色"
    ws["A3"].font = head_font
    ws["A3"].fill = PatternFill("solid", fgColor=c_header_bg)
    ws["B3"] = "意味"
    ws["B3"].font = head_font
    ws["B3"].fill = PatternFill("solid", fgColor=c_header_bg)

    color_rows = [
        (c_unknown, "groupId を特定できなかった行 (要手動補正)"),
        (c_inferred, "座標の一部をファイル名/マッピング/照合で推定・補完した行"),
        (c_band, "通常行 (交互の縞模様)"),
    ]
    r = 4
    for color, desc in color_rows:
        cA = ws.cell(row=r, column=1, value="")
        cA.fill = PatternFill("solid", fgColor=color)
        cA.border = border
        cB = ws.cell(row=r, column=2, value=desc)
        cB.font = body_font
        cB.alignment = Alignment(vertical="center", wrap_text=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="カラム").font = head_font
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=c_header_bg)
    ws.cell(row=r, column=1).font = head_font
    ws.cell(row=r, column=2, value="説明").font = head_font
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=c_header_bg)
    for cc in ("A", "B"):
        ws["%s%d" % (cc, r)].font = head_font
    r += 1

    col_desc = [
        ("groupId / artifactId / version",
         "Maven 座標。pom.properties→pom.xml→MANIFEST→既知マッピング→"
         "Maven Central 照合→ファイル名推定 の順に解決。"),
        ("classifier", "sources / javadoc 等の分類子 (ファイル名から推定)。"),
        ("ライブラリ名 / 概要", "pom.xml の name/description、MANIFEST、内蔵説明から構築。"),
        ("座標の判定元", "どの情報源で座標を決定したか。'ファイル名推定' は最も不確実。"),
        ("推定", "座標の一部を推測で補った場合に '推定' と表示。"),
        ("取得 URL", "Maven 標準レイアウトで生成した取得先。online 時は実在確認済み。"),
        ("URL 確認",
         "online: 実在確認 OK / 全リポジトリで未検出、offline: 未確認、"
         "座標未確定時: 生成不可。"),
        ("SHA-1", "ファイルの SHA-1。--online の Maven Central 照合キーにも使用。"),
    ]
    for label, desc in col_desc:
        ws.cell(row=r, column=1, value=label).font = body_font
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top",
                                                       wrap_text=True)
        cB = ws.cell(row=r, column=2, value=desc)
        cB.font = body_font
        cB.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1


# ----------------------------------------------------------------------------
# エントリポイント
# ----------------------------------------------------------------------------
def generate(scan_dir: str, output: str, online: bool, repos: List[str],
             timeout: int, include_old: bool,
             logger: logging.Logger) -> int:
    if not output.lower().endswith(".xlsx"):
        output = output + ".xlsx"
    reports = scan_jars(scan_dir, include_old, online, repos, timeout, logger)
    if not reports:
        logger.warning("JAR が見つかりませんでした: %s", scan_dir)
    write_xlsx(reports, output, scan_dir, online, logger)
    n_unknown = sum(1 for r in reports if r.groupId == UNKNOWN)
    if n_unknown:
        logger.warning("groupId 未確定が %d 件あります (Excel 上で橙色表示)。",
                       n_unknown)
    return 0


def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="jar_report.py",
        description="ディレクトリ配下の JAR を走査し Excel レポートを生成する")
    p.add_argument("--scan-dir", required=True, help="走査対象ディレクトリ")
    p.add_argument("--output", required=True, help="出力 xlsx パス")
    p.add_argument("--online", action="store_true",
                   help="Maven Central 照合で座標・概要を補完し URL を実在確認する")
    p.add_argument("--repo", action="append", default=None,
                   help="リポジトリ base URL (複数可、優先順)")
    p.add_argument("--timeout", type=int, default=30)
    p.add_argument("--include-old", action="store_true",
                   help="old/ 配下も対象にする (既定は除外)")
    p.add_argument("--log-file")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    logger = setup_logging(args.log_file)
    repos = args.repo if args.repo else DEFAULT_REPOS
    try:
        return generate(args.scan_dir, args.output, args.online, repos,
                        args.timeout, args.include_old, logger)
    except (ValueError, RuntimeError) as exc:
        logger.error("%s", exc)
        return 1
    except KeyboardInterrupt:
        logger.error("中断されました")
        return 130


if __name__ == "__main__":
    sys.exit(main())
